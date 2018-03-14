from __future__ import division
import openpyxl

def initialize(newSheet):
	newSheet['A1'] = "Year"
	newSheet['B1'] = "ParticipantID"
	newSheet['C1'] = "NeutralContext"

	newSheet['D1'] = "NeutralIdea1"
	newSheet['E1'] = "NeutralIdea2"
	newSheet['F1'] = "NeutralIdea3"
	newSheet['G1'] = "NeutralIdea4"
	newSheet['H1'] = "NeutralIdea5"
	newSheet['I1'] = "NeutralIdea6"
	newSheet['J1'] = "NeutralIdea7"
	newSheet['K1'] = "NeutralIdea8"
	newSheet['L1'] = "NeutralIdea9"
	newSheet['M1'] = "NeutralIdea10"
	newSheet['N1'] = "NeutralIdea11"
	newSheet['O1'] = "NeutralIdea12"
	newSheet['P1'] = "NeutralIdea13"
	newSheet['Q1'] = "NeutralIdea14"
	newSheet['R1'] = "NeutralIdea15"
	newSheet['S1'] = "NeutralIdea16"
	newSheet['T1'] = "NeutralAverage"

	newSheet['U1'] = "FramedContext"
	newSheet['V1'] = "FramingType"
	newSheet['W1'] = "FramedIdea1"
	newSheet['X1'] = "FramedIdea2"
	newSheet['Y1'] = "FramedIdea3"
	newSheet['Z1'] = "FramedIdea4"
	newSheet['AA1'] = "FramedIdea5"
	newSheet['AB1'] = "FramedIdea6"
	newSheet['AC1'] = "FramedIdea7"
	newSheet['AD1'] = "FramedIdea8"
	newSheet['AE1'] = "FramedIdea9"
	newSheet['AF1'] = "FramedIdea10"
	newSheet['AG1'] = "FramedIdea11"
	newSheet['AH1'] = "FramedIdea12"
	newSheet['AI1'] = "FramedIdea13"
	newSheet['AJ1'] = "FramedIdea14"
	newSheet['AK1'] = "FramedIdea15"
	newSheet['AL1'] = "FramedIdea16"
	newSheet['AM1'] = "FramedAverage"
	newSheet['AN1'] = "Shift(N-F)"

	return newSheet


def getIntervention(sheet, year, inColumn, row):
	intervention = sheet[inColumn + str(row)].value

	#If intervention is neutral
	if (intervention == "A") or (intervention == "N"):
		addValue(sheet, year, 0, row)
	#Else if intervention is adaptive framing
	elif (intervention == "B") or (intervention == "MI"):
		addValue(sheet, year, 1, row)
	#Else if intervention is innovative framing
	elif (intervention == "C") or (intervention == "MR"):
		addValue(sheet, year, 2, row)


def addValue(sheet, year, intervention, row):
	participantID = int(sheet['G' + str(row)].value)
	ideaCoding = int(sheet['AN' + str(row)].value)
	context = str(sheet['D' + str(row)].value)

	#If we have not seen an idea from this participant yet
	if participantID not in year[intervention]:
		year[intervention][participantID] = [context]
	#Add participant's idea code to their list
	year[intervention][participantID].append(ideaCoding)



def main():
	wb = openpyxl.load_workbook('IF-Data-6-Ideas-v09-with-PR-Codes-v2.xlsx')
	sheet = wb['Data']

	#Dictionaries for neutral, adaptive, and innovative ideas for each year
	yr1 = [{}, {}, {}]
	yr2 = [{}, {}, {}]
	yr3 = [{}, {}, {}]
	yr4 = [{}, {}, {}]

	dictionaries = [(1,yr1), (2,yr2), (3,yr3), (4,yr4)]

	yr1IDs = [1, 2, 3]
	yr1Values = ["Framing", "Neutral", "AllInterventions"]

	for row in range(2, sheet.max_row + 1):
		groupID = sheet['B' + str(row)].value
		ideaCoding = sheet['AN' + str(row)].value
		activityType = sheet['O' + str(row)].value
		studyDesign = sheet['AD' + str(row)].value

		try:
			#For year 1 data
			if ((groupID in yr1IDs) and (ideaCoding != '') and (activityType in yr1Values) and (studyDesign in yr1Values)):
				getIntervention(sheet, yr1, 'Q', row)

			#For year 2 data
			elif (groupID == 4) and (ideaCoding != ''):
				getIntervention(sheet, yr2, 'Q', row)

			#For year 3 data
			elif (groupID == 5) and (ideaCoding != ''):
				getIntervention(sheet, yr3, 'Q', row)

			#For year 4 data
			elif (groupID == 6) and (ideaCoding != ''):
				getIntervention(sheet, yr4, 'H', row)

		except:
			continue


	workbook = openpyxl.Workbook()
	newSheet = workbook.active

	newSheet = initialize(newSheet)

	# print(yr2[0])
	# print(yr2[1])

	problems = []
	row = 2
	for dictionary in dictionaries:
		for participants in dictionary[1][0]:
			if participants in dictionary[1][1]:
				framing = "Incremental"
				frameNum = 1
			elif participants in dictionary[1][2]:
				framing = "Radical"
				frameNum = 2
			else:
				framing = "None"
				frameNum = -1

			if frameNum == -1:	
				problems.append(participants)
				continue
			else:
				newSheet.cell(row=row, column=1).value = dictionary[0]
				newSheet.cell(row=row, column=2).value = participants
				newSheet.cell(row=row, column=3).value = dictionary[1][0][participants][0]

				neutralCount = 0
				column = 4
				for ideas in range(1, len(dictionary[1][0][participants])):
					newSheet.cell(row=row, column=column).value = dictionary[1][0][participants][ideas]
					if dictionary[1][0][participants][ideas] != 0:
						neutralCount += 1
					column += 1



				newSheet.cell(row=row, column=21).value = dictionary[1][frameNum][participants][0]
				newSheet.cell(row=row, column=22).value = framing

				framedCount = 0
				column = 23

				for ideas in range(1, len(dictionary[1][frameNum][participants])):
					newSheet.cell(row=row+1, column=column).value = dictionary[1][frameNum][participants][ideas]
					if dictionary[1][frameNum][participants][ideas] != 0:
						framedCount += 1
					column += 1


				neutralAvg = sum(dictionary[1][0][participants][1:])/neutralCount
				framedAvg = sum(dictionary[1][frameNum][participants][1:])/framedCount
				shift = neutralAvg - framedAvg
				newSheet.cell(row=row, column=20).value = round(neutralAvg, 3)
				newSheet.cell(row=row, column=39).value = round(framedAvg, 3)
				newSheet.cell(row=row, column=40).value = shift

				row += 1




	workbook.save("Framing-Outcomes-Individual-Shifts.xlsx")


if __name__ == '__main__':
    main()